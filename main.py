from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.filechooser import FileChooserListView
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.scrollview import ScrollView
from kivy.uix.popup import Popup
from kivy.core.window import Window
from kivy.core.clipboard import Clipboard
from kivy.clock import Clock
from kivy.utils import platform
from openpyxl import load_workbook
from collections import defaultdict
from datetime import datetime
import os
import shutil

# ================= ANDROID PERMISSION =================
if platform == "android":
    from android.permissions import request_permissions, Permission
    from android.storage import primary_external_storage_path
    request_permissions([
        Permission.READ_EXTERNAL_STORAGE,
        Permission.WRITE_EXTERNAL_STORAGE
    ])


class ClickableLabel(Label):
    def on_touch_down(self, touch):
        if self.collide_point(*touch.pos):
            if self.text.strip():
                Clipboard.copy(self.text)
                App.get_running_app().show_notif("Hasil berhasil dicopy!")
            return True
        return super().on_touch_down(touch)


class RekapApp(App):

    def build(self):
        Window.clearcolor = (0.05, 0.05, 0.07, 1)

        self.layout = BoxLayout(orientation='vertical', padding=15, spacing=15)

        header = Label(
            text="[b]REKAP DATA VOUCHER RUIJIE PRO[/b]",
            markup=True,
            font_size=26,
            size_hint=(1, 0.1),
            color=(1, 1, 1, 1)
        )
        self.layout.add_widget(header)

        self.result_label = ClickableLabel(
            text="Upload file excel",
            size_hint_y=None,
            markup=True,
            color=(1, 1, 1, 1)
        )
        self.result_label.bind(texture_size=self.result_label.setter('size'))

        scroll = ScrollView(size_hint=(1, 0.55))
        scroll.add_widget(self.result_label)
        self.layout.add_widget(scroll)

        self.btn_share = Button(
            text="Share Hasil Rekap",
            size_hint=(1, 0.1),
            background_normal="",
            background_color=(0.1, 0.5, 0.3, 1)
        )
        self.btn_share.bind(on_press=self.share_hasil)
        self.layout.add_widget(self.btn_share)

        btn_file = Button(
            text="Pilih File Excel",
            size_hint=(1, 0.1),
            background_normal="",
            background_color=(0.1, 0.3, 0.7, 1)
        )
        btn_file.bind(on_press=self.buka_file)
        self.layout.add_widget(btn_file)

        self.notif = Label(
            text="",
            size_hint=(1, 0.05),
            color=(1, 1, 0, 1)
        )
        self.layout.add_widget(self.notif)

        return self.layout

    # ================= NOTIF =================
    def show_notif(self, text):
        self.notif.text = text
        Clock.schedule_once(self.clear_notif, 2)

    def clear_notif(self, dt):
        self.notif.text = ""

    # ================= FILE CHOOSER =================
    def buka_file(self, instance):

        if platform == "android":
            storage_path = primary_external_storage_path()
        else:
            storage_path = "/"

        content = FileChooserListView(
            path=storage_path,
            filters=["*.xlsx", "*.xls"]
        )

        popup = Popup(
            title="Pilih File Excel",
            content=content,
            size_hint=(0.95, 0.95)
        )

        content.bind(
            on_submit=lambda chooser, selection, touch:
            self.proses_file(selection, popup)
        )

        popup.open()

    # ================= PROSES FILE =================
    def proses_file(self, selection, popup):
        if not selection:
            return

        file_path = selection[0]
        popup.dismiss()

        try:
            # Copy ke folder app dulu (anti crash android 11+)
            app_path = self.user_data_dir
            new_path = os.path.join(app_path, "temp.xlsx")
            shutil.copy(file_path, new_path)

            wb = load_workbook(new_path)
            sheet = wb.active

        except Exception:
            self.result_label.text = "[color=ff4444]Gagal membuka file![/color]"
            return

        rekap_detail = defaultdict(lambda: {"jumlah": 0, "total": 0})
        rekap_tanggal = defaultdict(int)

        header = [cell.value for cell in sheet[1]]

        try:
            kolom_grup = header.index("Grup pengguna")
            kolom_harga = header.index("Harga")
            kolom_tanggal = header.index("Diaktifkan di")
        except ValueError:
            self.result_label.text = "[color=ff4444]Header tidak sesuai![/color]"
            return

        for row in sheet.iter_rows(min_row=2, values_only=True):
            grup = row[kolom_grup]
            harga = row[kolom_harga]
            tanggal_full = row[kolom_tanggal]

            if grup and harga and tanggal_full:

                if isinstance(tanggal_full, datetime):
                    tanggal = tanggal_full.strftime("%Y/%m/%d")
                else:
                    tanggal = str(tanggal_full).split(" ")[0]

                try:
                    harga_int = int(float(harga))
                except:
                    continue

                key = (tanggal, grup)
                rekap_detail[key]["jumlah"] += 1
                rekap_detail[key]["total"] += harga_int
                rekap_tanggal[tanggal] += harga_int

        hasil = "[b]===== HASIL REKAP =====[/b]\n\n"
        tanggal_terakhir = None

        for (tanggal, grup), data in sorted(rekap_detail.items()):

            if tanggal_terakhir and tanggal != tanggal_terakhir:
                hasil += f"[color=FFA500]>>> TOTAL {tanggal_terakhir} : Rp {rekap_tanggal[tanggal_terakhir]}[/color]\n\n"

            if tanggal != tanggal_terakhir:
                hasil += f"\n[b]Tanggal : {tanggal}[/b]\n"

            hasil += f"  Grup   : {grup}\n"
            hasil += f"  Jumlah : {data['jumlah']}\n"
            hasil += f"  Total  : Rp {data['total']}\n\n"

            tanggal_terakhir = tanggal

        if tanggal_terakhir:
            hasil += f"[color=FFA500]>>> TOTAL {tanggal_terakhir} : Rp {rekap_tanggal[tanggal_terakhir]}[/color]\n"

        self.hasil_text = hasil
        self.result_label.text = hasil

    # ================= SHARE =================
    def share_hasil(self, instance):

        if not hasattr(self, "hasil_text"):
            self.show_notif("Belum ada hasil!")
            return

        if platform == "android":
            base_path = primary_external_storage_path()
            path = os.path.join(base_path, "hasil_rekap.txt")
        else:
            path = "hasil_rekap.txt"

        with open(path, "w", encoding="utf-8") as f:
            f.write(self.hasil_text)

        self.show_notif("File berhasil dibuat!")


if __name__ == "__main__":
    RekapApp().run()
