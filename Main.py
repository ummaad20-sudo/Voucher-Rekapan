from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.filechooser import FileChooserListView
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.scrollview import ScrollView
from kivy.uix.popup import Popup
from kivy.uix.floatlayout import FloatLayout
from kivy.graphics import Rectangle, Color, Ellipse, Line, RoundedRectangle
from kivy.graphics.texture import Texture
from kivy.core.window import Window
from kivy.clock import Clock
from kivy.utils import platform
from collections import defaultdict
from datetime import datetime
from openpyxl import load_workbook
import os
import random

# ===============================
# PERMISSION ANDROID (AMAN)
# ===============================
if platform == "android":
    try:
        from android.permissions import request_permissions, Permission
        request_permissions([
            Permission.READ_EXTERNAL_STORAGE,
            Permission.WRITE_EXTERNAL_STORAGE,
            Permission.INTERNET,
        ])
    except Exception:
        pass


# ===============================
# CLICK LABEL
# ===============================
class ClickableLabel(Label):
    def on_touch_down(self, touch):
        if self.collide_point(*touch.pos):
            if self.text.strip():
                from kivy.core.clipboard import Clipboard
                Clipboard.copy(self.text)
                App.get_running_app().show_notif("Hasil berhasil dicopy!")
            return True
        return super().on_touch_down(touch)


# ===============================
# GLASS CARD (STYLE TETAP)
# ===============================
class GlassCard(FloatLayout):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        with self.canvas.before:
            Color(0, 0, 0, 0.25)
            self.shadow = RoundedRectangle(radius=[25])

            Color(1, 1, 1, 0.10)
            self.bg = RoundedRectangle(radius=[25])

            Color(1, 1, 1, 0.35)
            self.border = Line(width=1.2)

        self.bind(pos=self.update_rect, size=self.update_rect)

    def update_rect(self, *args):
        self.shadow.pos = (self.x, self.y - 6)
        self.shadow.size = self.size
        self.bg.pos = self.pos
        self.bg.size = self.size
        self.border.rounded_rectangle = (
            self.x, self.y, self.width, self.height, 25
        )


# ===============================
# MAIN APP
# ===============================
class RekapApp(App):

    def rupiah(self, angka):
        try:
            return f"{int(angka):,}".replace(",", ".")
        except Exception:
            return str(angka)

    # ---------- BACKGROUND ----------
    def create_gradient(self, width, height):
        texture = Texture.create(size=(1, height), colorfmt='rgba')
        buf = []

        for y in range(height):
            t = y / float(height)
            r = int((1 - t) * 255 + t * 0)
            g = int((1 - t) * 120 + t * 80)
            b = int((1 - t) * 0 + t * 255)
            buf.extend([r, g, b, 255])

        texture.blit_buffer(bytes(buf), colorfmt='rgba', bufferfmt='ubyte')
        texture.wrap = 'repeat'
        texture.uvsize = (width, -1)
        return texture

    # ===============================
    def build(self):
        self.hasil_text = ""

        root = FloatLayout()

        self.layout = BoxLayout(
            orientation='vertical',
            padding=15,
            spacing=15,
            size_hint=(1, 1)
        )

        # ===== BACKGROUND =====
        with self.layout.canvas.before:
            self.gradient_texture = self.create_gradient(
                Window.width,
                Window.height
            )

            self.bg = Rectangle(
                texture=self.gradient_texture,
                size=Window.size,
                pos=(0, 0)
            )

            Color(0, 0, 0, 0.15)
            self.overlay = Rectangle(size=Window.size, pos=(0, 0))

            for _ in range(12):
                Color(1, 1, 1, 0.05)
                Ellipse(
                    pos=(random.randint(0, Window.width),
                         random.randint(0, Window.height)),
                    size=(random.randint(80, 200),
                          random.randint(80, 200))
                )

        self.layout.bind(size=self.update_bg)
        root.add_widget(self.layout)

        # ===== GLASS CARD =====
        card = GlassCard(
            size_hint=(0.92, 0.92),
            pos_hint={"center_x": 0.5, "center_y": 0.5}
        )

        content = BoxLayout(
            orientation='vertical',
            padding=20,
            spacing=15
        )

        # HEADER
        self.header = Label(
            text="[b]REKAPAN VOUCHER PENJUALAN[/b]",
            markup=True,
            font_size=35,
            size_hint=(1, 0.12),
            color=(1, 1, 1, 1)
        )
        content.add_widget(self.header)

        # HASIL
        self.result_label = ClickableLabel(
            text="MASUKAN FILE PENJUALAN",
            markup=True,
            size_hint_y=None,
            color=(1, 1, 1, 1)
        )
        self.result_label.bind(texture_size=self.result_label.setter('size'))

        scroll = ScrollView(size_hint=(1, 0.6))
        scroll.add_widget(self.result_label)
        content.add_widget(scroll)

        # SHARE
        self.btn_share = Button(
            text="SHARE HASIL REKAPAN",
            size_hint=(1, 0.12),
            background_color=(0, 0, 0, 0),
            color=(1, 1, 1, 1)
        )
        self.btn_share.bind(on_press=self.share_hasil)
        content.add_widget(self.btn_share)

        # FILE
        self.btn_file = Button(
            text="PILIH FILE EXCEL",
            size_hint=(1, 0.12),
            background_color=(0, 0, 0, 0),
            color=(1, 1, 1, 1)
        )
        self.btn_file.bind(on_press=self.buka_file)
        content.add_widget(self.btn_file)

        # NOTIF
        self.notif = Label(
            text="",
            size_hint=(1, 0.08),
            color=(1, 1, 0, 1)
        )
        content.add_widget(self.notif)

        # CREATOR
        self.creator = Label(
            text="creator by Jun © 2026",
            size_hint=(1, 0.08),
            font_size=25,
            color=(0.9, 0.9, 0.9, 1)
        )
        content.add_widget(self.creator)

        card.add_widget(content)
        root.add_widget(card)

        return root

    # ===============================
    def update_bg(self, *args):
        self.bg.size = Window.size
        self.overlay.size = Window.size

    def show_notif(self, text):
        self.notif.text = text
        Clock.schedule_once(lambda dt: self.clear_notif(), 2)

    def clear_notif(self, *args):
        self.notif.text = ""

    # ===============================
    def buka_file(self, instance):
        start_path = "/storage/emulated/0/" if platform == "android" else os.getcwd()

        chooser = FileChooserListView(
            path=start_path,
            filters=["*.xlsx"]
        )

        popup = Popup(
            title="Pilih File Excel",
            content=chooser,
            size_hint=(0.95, 0.95)
        )

        chooser.bind(
            on_submit=lambda ch, sel, touch:
            self.proses_file(sel, popup)
        )

        popup.open()

    # ===============================
    def proses_file(self, selection, popup):
        if not selection:
            return

        file_path = selection[0]
        popup.dismiss()

        try:
            wb = load_workbook(file_path, data_only=True)
            sheet = wb.active
        except Exception:
            self.result_label.text = "[color=ff4444]Gagal membuka file![/color]"
            return

        rekap_detail = defaultdict(lambda: {"jumlah": 0, "total": 0})
        rekap_tanggal = defaultdict(int)

        header = [str(c.value).strip() for c in sheet[1]]

        try:
            kolom_grup = header.index("Grup pengguna")
            kolom_harga = header.index("Harga")
            kolom_tanggal = header.index("Diaktifkan di")
        except ValueError:
            self.result_label.text = "[color=ff4444]Header tidak sesuai![/color]"
            return

        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                grup = row[kolom_grup]
                harga = row[kolom_harga]
                tanggal_full = row[kolom_tanggal]

                if not (grup and harga and tanggal_full):
                    continue

                if isinstance(tanggal_full, datetime):
                    tanggal = tanggal_full.strftime("%Y/%m/%d")
                else:
                    tanggal = str(tanggal_full).split(" ")[0]

                harga_int = int(float(harga))

                key = (tanggal, grup)
                rekap_detail[key]["jumlah"] += 1
                rekap_detail[key]["total"] += harga_int
                rekap_tanggal[tanggal] += harga_int
            except Exception:
                continue

        hasil = "[b]===== HASIL REKAP =====[/b]\n\n"
        tanggal_terakhir = None

        for (tanggal, grup), data in sorted(rekap_detail.items()):
            if tanggal_terakhir and tanggal != tanggal_terakhir:
                hasil += f"[color=FFA500]>>> TOTAL {tanggal_terakhir} : Rp {self.rupiah(rekap_tanggal[tanggal_terakhir])}[/color]\n"
                hasil += "-----------------------------\n"

            if tanggal != tanggal_terakhir:
                hasil += f"\n[b]Tanggal : {tanggal}[/b]\n"

            hasil += f"  Grup   : {grup}\n"
            hasil += f"  Jumlah : {data['jumlah']}\n"
            hasil += f"  Total  : Rp {self.rupiah(data['total'])}\n\n"

            tanggal_terakhir = tanggal

        if tanggal_terakhir:
            hasil += f"[color=FFA500]>>> TOTAL {tanggal_terakhir} : Rp {self.rupiah(rekap_tanggal[tanggal_terakhir])}[/color]\n"

        self.hasil_text = hasil
        self.result_label.text = hasil

    # ===============================
    def share_hasil(self, instance):
        if not self.hasil_text:
            self.show_notif("Belum ada hasil!")
            return

        try:
            file_path = "/storage/emulated/0/hasil_rekap.txt"

            with open(file_path, "w", encoding="utf-8") as f:
                f.write(self.hasil_text)

            if platform == "android":
                try:
                    from plyer import share
                    share.share(
                        title="Hasil Rekap",
                        text="File hasil rekap",
                        filepath=file_path
                    )
                except Exception as e:
                    self.show_notif(f"Gagal share: {e}")
                    return

            self.show_notif("File berhasil dibuat & dishare!")

        except Exception as e:
            self.show_notif(f"Gagal membuat file: {e}")


if __name__ == "__main__":
    RekapApp().run()
